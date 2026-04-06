import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from apps.candidates.models import Candidate
from apps.events.services import ScoringService


def generate_excel(candidate: Candidate) -> bytes:
    """
    Генерирует Excel отчёт по кандидату.
    Возвращает байты xlsx файла.
    """
    wb = Workbook()

    _build_summary_sheet(wb, candidate)
    _build_events_sheet(wb, candidate)
    _build_status_sheet(wb, candidate)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _header_style() -> dict:
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="534AB7"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            bottom=Side(style="thin", color="D3D1C7"),
        ),
    }


def _apply_style(cell, **kwargs) -> None:
    for attr, value in kwargs.items():
        setattr(cell, attr, value)


def _build_summary_sheet(wb: Workbook, candidate: Candidate) -> None:
    """Лист с основной информацией."""
    ws = wb.active
    ws.title = "Сводка"

    score = ScoringService.calculate(candidate)

    # Заголовок
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = f"Жизнь БРО — Отчёт: {candidate.name}"
    _apply_style(
        title_cell,
        font=Font(bold=True, size=14, color="534AB7"),
        alignment=Alignment(horizontal="center"),
    )
    ws.row_dimensions[1].height = 30

    # Данные
    rows = [
        ("Имя", candidate.name),
        ("Возраст", candidate.age or "—"),
        ("Познакомились", candidate.met_at or "—"),
        (
            "Тип привязанности",
            candidate.get_ai_attachment_type_display() or "—",
        ),
        ("Событий", candidate.events.count()),
        ("Hard Stop", "Да" if candidate.hard_stop_triggered else "Нет"),
        (
            "Итоговый скор",
            f"{score}/100" if score is not None else "Нет данных",
        ),
        ("Добавлен", candidate.created_at.strftime("%d.%m.%Y")),
        ("Дата отчёта", datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        label_cell = ws.cell(row=i, column=1, value=label)
        value_cell = ws.cell(row=i, column=2, value=value)
        _apply_style(
            label_cell,
            font=Font(bold=True),
            fill=PatternFill("solid", fgColor="F1EFE8"),
        )
        if i % 2 == 0:
            _apply_style(
                value_cell,
                fill=PatternFill("solid", fgColor="FFFFFF"),
            )

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40

    # Watermark
    ws.cell(
        row=len(rows) + 4,
        column=1,
        value="Сгенерировано ботом @zhizn_bro_bot",
    ).font = Font(italic=True, color="888780", size=9)


def _build_events_sheet(wb: Workbook, candidate: Candidate) -> None:
    """Лист с событиями и оценками."""
    ws = wb.create_sheet("События")

    headers = [
        "Дата",
        "Событие",
        "Критерий",
        "Балл ИИ",
        "Балл итог",
        "Подтверждено",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style = _header_style()
        _apply_style(cell, **style)

    row = 2
    for event in candidate.events.prefetch_related("scores__criterion").all():
        scores = event.scores.all()
        if scores:
            for score in scores:
                ws.cell(
                    row=row,
                    column=1,
                    value=event.created_at.strftime("%d.%m.%Y"),
                )
                ws.cell(row=row, column=2, value=event.raw_text[:100])
                ws.cell(row=row, column=3, value=score.criterion.name)
                ws.cell(row=row, column=4, value=score.ai_score)
                ws.cell(row=row, column=5, value=score.final_score)
                ws.cell(
                    row=row,
                    column=6,
                    value="Да" if score.is_confirmed else "Нет",
                )
                row += 1
        else:
            ws.cell(
                row=row, column=1, value=event.created_at.strftime("%d.%m.%Y")
            )
            ws.cell(row=row, column=2, value=event.raw_text[:100])
            row += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14


def _build_status_sheet(wb: Workbook, candidate: Candidate) -> None:
    """Лист с историей статусов."""
    ws = wb.create_sheet("Статусы")

    headers = ["Статус", "Дата", "Примечание"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style = _header_style()
        _apply_style(cell, **style)

    for row, sh in enumerate(candidate.status_history.all(), start=2):
        ws.cell(row=row, column=1, value=sh.get_status_display())
        ws.cell(row=row, column=2, value=sh.started_at.strftime("%d.%m.%Y"))
        ws.cell(row=row, column=3, value=sh.note or "—")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
